from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
from PIL import Image

def drawPictures(IMAGE_FILE):
    try:
        img = Image.open(IMAGE_FILE)
    except:
        return -1
    print('image\'s type is %s' % img.format)

    IMAGE_MAX_SIZE=1000

    img_width, img_height = img.size
    # 修正图片尺寸
    img_width, img_height = img.size
    if img_width > IMAGE_MAX_SIZE:
        img_height = int(img_height * IMAGE_MAX_SIZE / img_width)
        img_width = IMAGE_MAX_SIZE

    if img_height > IMAGE_MAX_SIZE:
        img_width = int(img_width * IMAGE_MAX_SIZE / img_height)
        img_height = IMAGE_MAX_SIZE

    img.thumbnail((img_width, img_height))  # 缩小图片尺寸
    print('image\'s new size is (%s, %s)' % (img_width, img_height))
    
    fName=IMAGE_FILE.split('/')[-1].split('.')[0]
    if len(fName)>15:
        fName=fName[-15:]
    print(fName)
    # 由于excel对单元格填充色有要求，所以用以下两行代码把图片转换为8位色值
    img = img.convert('P')
    img = img.convert('RGB')

    pix = img.load()

    workbook = Workbook()
    worksheet = workbook.active
    print('begin convert, please waiting...')

    # # 获取每个像素的色值，并填充到单元格
    for row in range(1, img_height):
        for col in range(1, img_width):
            cell = worksheet.cell(column=col, row=row)
            point = pix[col-1, row-1]
            color = "FF%02X%02X%02X" % (point[0], point[1], point[2])
            cell.fill = PatternFill(patternType='solid', fgColor=Color(rgb=color))
        worksheet.row_dimensions[row].height = 6

    for col in range(1, img_width):
        worksheet.column_dimensions[get_column_letter(col)].width = 10/9

    # 保存生成的excel文件

    workbook.save(filename=fName+'.xlsx')
    print('Complete.')
    return 1